"""
Reconciliation tool for supplier Excel statement verification.

Supports two suppliers:
- 新国线 (supplier_1): Itinerary-level reconciliation
  - Detect:    header contains "供应商账号"
  - Match key: 订单ID (col 0) → lulu_order_itinerary.open_order_no
  - Compare:   票数 → COUNT(lulu_ticket)
               总金额 → SUM(lulu_order_itinerary.actual_amt)  支持双程单
               退款金额 → SUM(lulu_order_itinerary.refund_amt)
               订单状态 → MIN(lulu_order_itinerary.state)
  - Skip:      已取消订单跳过所有金额核对
  - Finance:   佣金率 8%；输出有效票数、总金额、佣金、应结款

- 车盈网 (supplier_2): Ticket-level reconciliation
  - Detect:    header contains "车盈网条码"
  - Match key: 车盈网条码 (col 36) → lulu_ticket.open_ticket_no
  - Compare:   支付金额 (col 61) → lulu_ticket.pay_amt
               车票状态 (col 17) → lulu_ticket.state
  - Finance:   空港票源（售票渠道2含"白云机场"）佣金率 3%；其余 10%
               按空港/非空港分别输出售票金额、退票金额、有效票金额、有效票数、代售佣金
"""

import json
import os
from typing import Any, Dict, List, Optional, Tuple

from agent.tools.base_tool import BaseTool, ToolResult
from common.log import logger
from config import conf

# ── Supplier detection ───────────────────────────────────────────────────────

_SUPPLIER_1_MARKER = "供应商账号"   # unique column in 新国线
_SUPPLIER_2_MARKER = "车盈网条码"   # unique column in 车盈网

# ── Supplier 1 column indices (0-based) ─────────────────────────────────────

S1_COL_ORDER_ID   = 0
S1_COL_TICKET_CNT = 5
S1_COL_CHILD_CNT  = 6   # 儿童票数，计入总票数
S1_COL_TOTAL_AMT  = 14
S1_COL_REFUND_AMT = 15
S1_COL_STATUS     = 18

S1_COMMISSION_RATE = 0.08

# supplier status → expected lulu_order_itinerary.state values
S1_STATUS_MAP: Dict[str, set] = {
    "普通":   {100, 200, 210, 260, 300},
    "已取消": {400, 450, 500},
}

# ── Supplier 2 column indices (0-based) ─────────────────────────────────────

S2_COL_STATUS    = 17
S2_COL_PAY_AMT   = 61
S2_COL_BARCODE   = 36   # 车盈网条码 → lulu_ticket.open_ticket_no
S2_COL_ORDER_NO  = 27   # 订单号，用于报告展示
S2_COL_PASSENGER = 32   # 乘车人姓名，用于报告展示

S2_AIRPORT_COL_NAME  = "售票渠道2"   # 列名，用于识别空港票源
S2_STATION_BC_COL_NAME = "车站条码" # 列名，车盈网条码找不到时的备用查询键
S2_AIRPORT_MARKER   = "白云机场"    # 列值含此关键字 → 空港票源
S2_AIRPORT_RATE     = 0.03
S2_NORMAL_RATE      = 0.10

# 空港票源：供应商状态 → lulu_order_itinerary.state 期望值（和 lulu_order 同枚举）
S2_AIRPORT_ITINERARY_STATUS_MAP: Dict[str, set] = {
    "valid":  {100, 200, 210, 260, 300},
    "refund": {400, 450, 500},
}

# supplier status labels
S2_VALID_STATUSES  = {"已检", "已售", "混检"}
S2_REFUND_STATUSES = {"已退"}

# expected lulu_ticket.state values
S2_STATUS_MAP: Dict[str, set] = {
    "valid":  {200, 300},
    "refund": {0, 400, 500},
}

# ── Misc ─────────────────────────────────────────────────────────────────────

_BATCH = 200        # DB IN-clause batch size
_AMT_TOLERANCE = 0.01  # float comparison tolerance


class ReconciliationTool(BaseTool):
    """对账供应商Excel账单，逐行比对订单/票据状态和金额是否与系统一致。"""

    name: str = "reconciliation"
    description: str = (
        "对账供应商Excel账单，逐行比对订单状态和金额是否与系统一致。"
        "支持新国线（行程级）和车盈网（票级）两种账单格式，自动识别。"
        "传入文件路径，返回差异明细、汇总统计和财务汇总。"
    )
    params: dict = {
        "type": "object",
        "properties": {
            "file_path": {
                "type": "string",
                "description": "供应商Excel账单的本地文件路径（.xlsx 或 .xls）"
            }
        },
        "required": ["file_path"]
    }

    def execute(self, args: Dict[str, Any]) -> ToolResult:
        file_path = args.get("file_path", "").strip()
        if not file_path:
            return ToolResult.fail("Error: file_path is required")

        if not os.path.exists(file_path):
            return ToolResult.fail(f"Error: 文件不存在: {file_path}")

        try:
            rows, headers = self._read_excel(file_path)
        except ImportError:
            return ToolResult.fail(
                "Error: openpyxl 未安装，请执行: pip install openpyxl"
            )
        except Exception as e:
            return ToolResult.fail(f"Error: 读取 Excel 失败: {e}")

        if not rows:
            return ToolResult.fail("Error: Excel 文件为空或无数据行")

        header_set = {str(h).strip() for h in headers if h}
        if _SUPPLIER_1_MARKER in header_set:
            return self._reconcile_supplier1(rows)
        elif _SUPPLIER_2_MARKER in header_set:
            return self._reconcile_supplier2(rows, headers)
        else:
            return ToolResult.fail(
                "无法识别供应商格式，请确认是新国线或车盈网的账单。"
            )

    # ── Excel reader ─────────────────────────────────────────────────────────

    @staticmethod
    def _read_excel(file_path: str) -> Tuple[List[tuple], tuple]:
        """读取 Excel，自动扫描所有 sheet，返回第一个含供应商标记列头的 sheet 数据。"""
        import openpyxl
        markers = {_SUPPLIER_1_MARKER, _SUPPLIER_2_MARKER}
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                all_rows = list(ws.iter_rows(values_only=True))
                if not all_rows:
                    continue
                header_set = {str(h).strip() for h in all_rows[0] if h}
                if markers & header_set:  # 找到至少一个供应商标记
                    headers = all_rows[0]
                    data_rows = [r for r in all_rows[1:] if any(c is not None for c in r)]
                    return data_rows, headers
        finally:
            wb.close()
        return [], ()

    # ── DB connection ─────────────────────────────────────────────────────────

    @staticmethod
    def _connect():
        import pymysql
        import pymysql.cursors
        return pymysql.connect(
            host=conf().get("mysql_host", "localhost"),
            port=int(conf().get("mysql_port", 3306)),
            user=conf().get("mysql_user", ""),
            password=conf().get("mysql_password", ""),
            database=conf().get("mysql_database", "") or None,
            charset="utf8mb4",
            cursorclass=pymysql.cursors.DictCursor,
            connect_timeout=10,
        )

    # ── Supplier 1：新国线（行程级） ──────────────────────────────────────────

    def _reconcile_supplier1(self, rows: List[tuple]) -> ToolResult:
        order_ids = [
            str(r[S1_COL_ORDER_ID]).strip()
            for r in rows
            if r[S1_COL_ORDER_ID]
        ]
        if not order_ids:
            return ToolResult.fail("账单中未找到有效的订单ID")

        db_ticket_counts  = self._query_ticket_counts_batch(order_ids)
        db_itinerary_amts = self._query_itinerary_amounts_batch(order_ids)

        issues = []
        not_found = []
        matched = 0
        total_sales_amount    = 0.0
        total_valid_ticket_cnt = 0

        for row in rows:
            oid = str(row[S1_COL_ORDER_ID]).strip() if row[S1_COL_ORDER_ID] else None
            if not oid:
                continue

            if oid not in db_itinerary_amts:
                not_found.append(oid)
                continue

            # 已取消的订单：相当于不存在，跳过所有金额核对
            supplier_status = str(row[S1_COL_STATUS] or "").strip()
            if supplier_status == "已取消":
                matched += 1
                continue

            itinerary = db_itinerary_amts[oid]
            row_issues = []

            # 累计有效销售金额和票数
            supplier_amt = float(row[S1_COL_TOTAL_AMT] or 0)
            supplier_cnt = int(row[S1_COL_TICKET_CNT] or 0) + int(row[S1_COL_CHILD_CNT] or 0)
            total_sales_amount    += supplier_amt
            total_valid_ticket_cnt += supplier_cnt

            # 票数（成人票 + 儿童票）
            db_cnt = db_ticket_counts.get(oid, 0)
            if supplier_cnt != db_cnt:
                row_issues.append({
                    "field": "票数",
                    "supplier": supplier_cnt,
                    "db": db_cnt,
                })

            # 总金额 → SUM(lulu_order_itinerary.actual_amt)（支持双程单）
            db_itinerary_amt = itinerary["actual_amt"]
            if abs(supplier_amt - db_itinerary_amt) > _AMT_TOLERANCE:
                row_issues.append({
                    "field": "总金额",
                    "supplier": supplier_amt,
                    "db": db_itinerary_amt,
                })

            # 退款金额 → SUM(lulu_order_itinerary.refund_amt)
            supplier_refund = float(row[S1_COL_REFUND_AMT] or 0)
            db_refund = itinerary["refund_amt"]
            if abs(supplier_refund - db_refund) > _AMT_TOLERANCE:
                row_issues.append({
                    "field": "退款金额",
                    "supplier": supplier_refund,
                    "db": db_refund,
                })

            # 订单状态 → MIN(lulu_order_itinerary.state)
            db_state = itinerary["state"]
            expected = S1_STATUS_MAP.get(supplier_status)
            if expected is not None and db_state not in expected:
                row_issues.append({
                    "field": "订单状态",
                    "supplier": supplier_status,
                    "db": db_state,
                })

            if row_issues:
                issues.append({"order_no": oid, "issues": row_issues})
            else:
                matched += 1

        commission = round(total_sales_amount * S1_COMMISSION_RATE, 2)
        financial_summary = {
            "有效票数": total_valid_ticket_cnt,
            "总金额": round(total_sales_amount, 2),
            "佣金比例": f"{int(S1_COMMISSION_RATE * 100)}%",
            "佣金": commission,
            "应结款": round(total_sales_amount - commission, 2),
        }
        return self._build_result(
            "新国线", len(order_ids), matched, issues, not_found, financial_summary,
        )

    # ── Supplier 2：车盈网（票级/行程级混合） ────────────────────────────────

    def _reconcile_supplier2(self, rows: List[tuple], headers: tuple) -> ToolResult:
        # 找"售票渠道2"和"车站条码"列的索引
        airport_col:    Optional[int] = None
        station_bc_col: Optional[int] = None
        for idx, h in enumerate(headers):
            hs = str(h).strip() if h else ""
            if hs == S2_AIRPORT_COL_NAME:
                airport_col = idx
            elif hs == S2_STATION_BC_COL_NAME:
                station_bc_col = idx

        # 按空港/非空港分流
        airport_rows: List[tuple] = []
        normal_rows:  List[tuple] = []
        for row in rows:
            is_airport = (
                airport_col is not None
                and row[airport_col] is not None
                and S2_AIRPORT_MARKER in str(row[airport_col])
            )
            if is_airport:
                airport_rows.append(row)
            else:
                normal_rows.append(row)

        issues:    List[dict] = []
        not_found: List[str]  = []
        matched = 0
        airport_stats = {"sell": 0.0, "refund": 0.0, "valid_cnt": 0}
        normal_stats  = {"sell": 0.0, "refund": 0.0, "valid_cnt": 0}

        # ── 非空港：逐票对比 lulu_ticket ──────────────────────────────────────
        # 批量预查询：车盈网条码 + 车站条码（都映射到 open_ticket_no）
        barcodes = [str(r[S2_COL_BARCODE]).strip() for r in normal_rows if r[S2_COL_BARCODE]]
        station_barcodes = (
            [str(r[station_bc_col]).strip() for r in normal_rows
             if station_bc_col is not None and r[station_bc_col] and str(r[station_bc_col]).strip()]
            if station_bc_col is not None else []
        )
        all_lookup_keys = list(set(barcodes + station_barcodes))
        db_tickets = self._query_tickets_batch(all_lookup_keys) if all_lookup_keys else {}

        for row in normal_rows:
            bc = str(row[S2_COL_BARCODE]).strip() if row[S2_COL_BARCODE] else None
            if not bc:
                continue

            supplier_status = str(row[S2_COL_STATUS] or "").strip()
            supplier_amt    = float(row[S2_COL_PAY_AMT] or 0)

            # 先用车盈网条码查，找不到则 fallback 到车站条码
            db = db_tickets.get(bc)
            if db is None and station_bc_col is not None:
                station_bc = str(row[station_bc_col] or "").strip()
                if station_bc:
                    db = db_tickets.get(station_bc)

            if db is None:
                not_found.append(bc)
                continue

            db = db_tickets[bc]
            row_issues = []

            db_amt = float(db["pay_amt"] or 0)
            if abs(supplier_amt - db_amt) > _AMT_TOLERANCE:
                row_issues.append({"field": "支付金额", "supplier": supplier_amt, "db": db_amt})

            db_state = int(db["state"] or 0)
            if supplier_status in S2_VALID_STATUSES:
                expected = S2_STATUS_MAP["valid"]
            elif supplier_status in S2_REFUND_STATUSES:
                expected = S2_STATUS_MAP["refund"]
            else:
                expected = None
            if expected is not None and db_state not in expected:
                row_issues.append({"field": "车票状态", "supplier": supplier_status, "db": db_state})

            if row_issues:
                issues.append({
                    "order_no":  str(row[S2_COL_ORDER_NO] or bc),
                    "barcode":   bc,
                    "passenger": str(row[S2_COL_PASSENGER] or ""),
                    "issues":    row_issues,
                })
            else:
                matched += 1

            if supplier_status in S2_VALID_STATUSES:
                normal_stats["sell"]      += supplier_amt
                normal_stats["valid_cnt"] += 1
            elif supplier_status in S2_REFUND_STATUSES:
                normal_stats["refund"] += supplier_amt

        # ── 空港：按订单号分组，对比 lulu_order_itinerary ─────────────────────
        # 先按 order_no 分组
        airport_order_rows: Dict[str, List[tuple]] = {}
        for row in airport_rows:
            order_no = str(row[S2_COL_ORDER_NO] or "").strip()
            if not order_no:
                continue
            airport_order_rows.setdefault(order_no, []).append(row)

        db_itineraries = (
            self._query_itinerary_amounts_batch(list(airport_order_rows.keys()))
            if airport_order_rows else {}
        )

        for order_no, order_rows in airport_order_rows.items():
            valid_rows  = [r for r in order_rows if str(r[S2_COL_STATUS] or "").strip() in S2_VALID_STATUSES]
            refund_rows = [r for r in order_rows if str(r[S2_COL_STATUS] or "").strip() in S2_REFUND_STATUSES]
            valid_amt   = sum(float(r[S2_COL_PAY_AMT] or 0) for r in valid_rows)
            refund_amt  = sum(float(r[S2_COL_PAY_AMT] or 0) for r in refund_rows)

            if order_no not in db_itineraries:
                not_found.append(order_no)
                continue

            itinerary  = db_itineraries[order_no]
            row_issues = []

            # 有效票金额合计 vs lulu_order_itinerary.actual_amt
            db_actual_amt = itinerary["actual_amt"]
            if abs(valid_amt - db_actual_amt) > _AMT_TOLERANCE:
                row_issues.append({
                    "field":    "支付金额",
                    "supplier": valid_amt,
                    "db":       db_actual_amt,
                })

            # 状态：有效票为主，全退则看退款状态
            if valid_rows:
                rep_status = str(valid_rows[0][S2_COL_STATUS] or "").strip()
                expected_states = S2_AIRPORT_ITINERARY_STATUS_MAP["valid"]
            elif refund_rows:
                rep_status = str(refund_rows[0][S2_COL_STATUS] or "").strip()
                expected_states = S2_AIRPORT_ITINERARY_STATUS_MAP["refund"]
            else:
                rep_status, expected_states = "", None

            db_state = itinerary["state"]
            if expected_states is not None and db_state not in expected_states:
                row_issues.append({
                    "field":    "行程状态",
                    "supplier": rep_status,
                    "db":       db_state,
                })

            if row_issues:
                issues.append({
                    "order_no":     order_no,
                    "ticket_count": len(order_rows),
                    "issues":       row_issues,
                })
            else:
                matched += len(order_rows)

            airport_stats["sell"]      += valid_amt
            airport_stats["refund"]    += refund_amt
            airport_stats["valid_cnt"] += len(valid_rows)

        # ── 财务汇总 ──────────────────────────────────────────────────────────
        def _bucket_summary(bucket: dict, rate: float) -> dict:
            sell       = round(bucket["sell"], 2)
            refund     = round(bucket["refund"], 2)
            effective  = round(sell - refund, 2)
            commission = round(effective * rate, 2)
            return {
                "有效票金额":  effective,
                "退票金额":   refund,
                "有效票数":   bucket["valid_cnt"],
                "佣金比例":   f"{int(rate * 100)}%",
                "代售佣金":   commission,
            }

        financial_summary = {
            "空港票源":   _bucket_summary(airport_stats, S2_AIRPORT_RATE),
            "非空港票源":  _bucket_summary(normal_stats,  S2_NORMAL_RATE),
        }

        total = len(barcodes) + sum(len(v) for v in airport_order_rows.values())
        return self._build_result(
            "车盈网", total, matched, issues, not_found, financial_summary,
        )

    # ── DB queries ────────────────────────────────────────────────────────────

    def _query_ticket_counts_batch(self, order_ids: List[str]) -> Dict[str, int]:
        result: Dict[str, int] = {}
        conn = None
        try:
            conn = self._connect()
            with conn.cursor() as cur:
                for i in range(0, len(order_ids), _BATCH):
                    batch = order_ids[i:i + _BATCH]
                    ph = ",".join(["%s"] * len(batch))
                    cur.execute(
                        f"SELECT open_order_no, COUNT(*) AS cnt "
                        f"FROM lulu_ticket "
                        f"WHERE open_order_no IN ({ph}) AND del_flag=0 "
                        f"GROUP BY open_order_no",
                        batch,
                    )
                    for row in cur.fetchall():
                        result[row["open_order_no"]] = int(row["cnt"])
        except Exception as e:
            logger.error(f"[ReconciliationTool] Query ticket counts failed: {e}")
        finally:
            if conn:
                conn.close()
        return result

    def _query_itinerary_amounts_batch(self, order_ids: List[str]) -> Dict[str, dict]:
        """按 open_order_no 汇总 lulu_order_itinerary 的实际金额、退款金额和最小状态。"""
        result: Dict[str, dict] = {}
        conn = None
        try:
            conn = self._connect()
            with conn.cursor() as cur:
                for i in range(0, len(order_ids), _BATCH):
                    batch = order_ids[i:i + _BATCH]
                    ph = ",".join(["%s"] * len(batch))
                    cur.execute(
                        f"SELECT open_order_no, "
                        f"  SUM(actual_amt) AS actual_amt, "
                        f"  SUM(refund_amt) AS refund_amt, "
                        f"  MIN(state)      AS state "
                        f"FROM lulu_order_itinerary "
                        f"WHERE open_order_no IN ({ph}) AND del_flag=0 "
                        f"GROUP BY open_order_no",
                        batch,
                    )
                    for row in cur.fetchall():
                        result[row["open_order_no"]] = {
                            "actual_amt": float(row["actual_amt"] or 0),
                            "refund_amt": float(row["refund_amt"] or 0),
                            "state":      int(row["state"] or 0),
                        }
        except Exception as e:
            logger.error(f"[ReconciliationTool] Query itinerary amounts failed: {e}")
        finally:
            if conn:
                conn.close()
        return result

    def _query_tickets_batch(self, barcodes: List[str]) -> Dict[str, dict]:
        result: Dict[str, dict] = {}
        conn = None
        try:
            conn = self._connect()
            with conn.cursor() as cur:
                for i in range(0, len(barcodes), _BATCH):
                    batch = barcodes[i:i + _BATCH]
                    ph = ",".join(["%s"] * len(batch))
                    cur.execute(
                        f"SELECT open_ticket_no, pay_amt, state "
                        f"FROM lulu_ticket "
                        f"WHERE open_ticket_no IN ({ph}) AND del_flag=0",
                        batch,
                    )
                    for row in cur.fetchall():
                        result[row["open_ticket_no"]] = row
        except Exception as e:
            logger.error(f"[ReconciliationTool] Query tickets failed: {e}")
        finally:
            if conn:
                conn.close()
        return result

    # ── Result builder ────────────────────────────────────────────────────────

    @staticmethod
    def _build_result(
        supplier: str,
        total: int,
        matched: int,
        issues: List[dict],
        not_found: List[str],
        financial_summary: dict,
    ) -> ToolResult:
        result = {
            "supplier": supplier,
            "total": total,
            "matched": matched,
            "issue_count": len(issues),
            "not_found_count": len(not_found),
            "issues": issues,
            "not_found": not_found[:50],  # 防止 token 溢出
            "financial_summary": financial_summary,
        }
        logger.info(
            f"[ReconciliationTool] {supplier}: total={total}, matched={matched}, "
            f"issues={len(issues)}, not_found={len(not_found)}"
        )
        return ToolResult.success(json.dumps(result, ensure_ascii=False, indent=2))
